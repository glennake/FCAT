#!/usr/bin/env python3

__all__ = ["FCAT_UI", "FortiCRP", "FortiCareAPI"]
__author__ = "Glenn Akester (@glennake)"
__description__ = "Fortinet Contract Automation Tool (FCAT) is a Python application to assist \
with bulk parsing of Fortinet contract ZIP files and the registration \
of Fortinet products and services using Fortinet's FortiCare API."
__title__ = "Fortinet Contract Automation Tool (FCAT)"
__version_info__ = (0, 1, 0)
__version__ = ".".join(map(str, __version_info__))

# Global Imports

import csv
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
import json
import os
import platform
import re
import threading
import time
from traceback_with_variables import format_exc
from zipfile import ZipFile

thread_run = False


class FortiCareAPI:
    """A class to represent a FortiCare API connection instance.

    Parameters
    ----------
    username : str
        Username for FortiCare API authentication.
    password : str
        Password for FortiCare API authentication.
    """

    def __init__(self, username, password):
        """Constructs all the necessary attributes for the FortiCareAPI object.
        """

        import json
        import requests

        self.json = json
        self.requests = requests

        self.api_oauth = "https://customerapiauth.fortinet.com/api/v1/oauth/"
        self.api_base = "https://support.fortinet.com/ES/api/registration/v3/"
        self.api_user = username
        self.api_pass = password
        self.api_client_id = "assetmanagement"

        self.api_token = None
        self.api_refresh_token = None

        self.batch_reg_max = 10 # FortiCare API maximum entries in a bulk request set by Fortinet (default=10)

        self.api_limit_req_minute = 100 # FortiCare API calls per minute limit set by Fortinet (default=100)
        self.api_req_minute = 0

        self.api_limit_req_hour = 1000 # FortiCare API calls per hour limit set by Fortinet (default=1000)
        self.api_req_hour = 0

        self.api_limit_err_hour = 10 # FortiCare API errors per hour limit set by Fortinet (default=10)
        self.api_err_hour = 0

    def _check_rate_limits(self, req_len=1):
        if ( # Check if the error counter exceeds the error limit
            self.api_err_hour >= self.api_limit_err_hour
        ):  # TODO: Enhance as this does not actually check if all errors were in the last hour
            time.sleep(
                3600
            )  # TODO: Add some feedback here so the user knows we are pausing requests
            self.api_err_hour = 0
            self.api_req_hour = 0
            self.api_req_minute = 0
        elif ( # Check if the request counter exceeds the calls per hour limit
            self.api_req_hour + req_len
        ) >= self.api_limit_req_hour:  # TODO: Enhance as this does not  check if all requests were in the last hour
            time.sleep(
                3600
            )  # TODO: Add some feedback here so the user knows we are pausing requests
            self.api_err_hour = 0
            self.api_req_hour = 0
            self.api_req_minute = 0
        elif ( # Check if the request counter exceeds the calls per minute limit
            self.api_req_minute + req_len
        ) >= self.api_limit_req_minute:  # TODO: Enhance as this does not check if all requests were in the last minute
            time.sleep(
                60
            )  # TODO: Add some feedback here so the user knows we are pausing requests
            self.api_req_minute = 0

    def _increment_req_limit_counters(self, increment=1):
        self.api_req_minute += increment
        self.api_req_hour += increment

    def _increment_err_limit_counters(self, increment=1):
        self.api_err_hour += increment

    def _send_api_request(self, url, data, headers={}, method="post"):
        self._check_rate_limits()

        if self.api_token:
            headers = { # any additional headers passed in are currently ignored and replaced
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(self.api_token),
            }
        else:
            headers = { # any additional headers passed in are currently ignored and replaced
                "Content-Type": "application/json",
            }

        data_str = self.json.dumps(data)

        try:
            r = self.requests.request(method, url, headers=headers, data=data_str)
            r_sc = r.status_code
            r_json = self.json.loads(r.text)
        except:
            r = None
            r_sc = 0
            r_json = None

        self._increment_req_limit_counters()

        if r_json and r_json["status"] != 0:
            self._increment_err_limit_counters()

        return r, r_sc, r_json

    def get_oauth_token(self):
        """Authenticates to the FortiCare API and stores the received bearer token.

        Returns
        -------
        bool
            True if successful, False if unsuccessful.
        """

        self.api_token = None

        url = self.api_oauth + "token/"

        data = {
            "username": self.api_user,
            "password": self.api_pass,
            "client_id": self.api_client_id,
            "grant_type": "password",
        }

        r, r_sc, r_json = self._send_api_request(url, data=data)

        if r and r_sc == 200:
            self.api_token = r_json["access_token"]
            self.api_refresh_token = r_json["refresh_token"]

        if self.api_token and self.api_refresh_token:
            return True
        else:
            return False

    def revoke_oauth_token(self):
        """Unuthenticates from the FortiCare API.

        Returns
        -------
        bool
            True if successful, False if unsuccessful.
        """

        url = self.api_oauth + "revoke_token/"

        data = {"client_id": self.api_client_id, "token": self.api_token,}

        r, r_sc, r_json = self._send_api_request(url, data=data)

        if r and r_sc == 200:
            self.api_token = None
            self.api_refresh_token = None

        if self.api_token and self.api_refresh_token:
            return False
        else:
            return True
        
    def get_asset_details(self, assets: list):
        """Get Fortinet asset details using the FortiCare API from a list of serial numbers.

        Parameters
        ----------
        assets : list
            List containing serial numbers::
            
                [
                    "S424ENTF10001234",
                    "S424ENTF10001235",
                    "S424ENTF10001236",
                ]

        Yields
        ------
        dict
            Dictionary containing information returned by the product details API call::

            {
                "serialNumber": "S424ENTF10001234",
                "productModel": "FortiSwitch 424E",
                "description": "MYFSW1"
            }
        """

        url = self.api_base + "products/details"

        for a in assets:
            data = {"serialNumber": a}

            r, r_sc, r_json = self._send_api_request(url, data=data)

            if r_json and r_sc in [200]:
                if "assetDetails" in r_json and r_json["assetDetails"]:
                    yield {
                        "serialNumber": r_json["assetDetails"]["serialNumber"],
                        "productModel": r_json["assetDetails"]["productModel"],
                        "description": r_json["assetDetails"]["description"],
                    }

                else:
                    yield {
                        "serialNumber": None,
                        "productModel": None,
                        "description": None,
                        "status": None,
                        "message": None,
                        "gbl_status": r_json["status"],
                        "gbl_message": r_json["message"],
                    }

            else:
                yield {
                    "serialNumber": None,
                    "productModel": None,
                    "description": None,
                    "status": None,
                    "message": None,
                    "gbl_status": 2,
                    "gbl_message": r_json["message"],
                }


    def register_assets(self, assets):
        """Registers Fortinet assets using the FortiCare API from a list of dictionaries containing asset information.

        Parameters
        ----------
        assets : list
            List of dictionaries containing asset data::
            
                [
                    {
                        'serialNumber': "S424ENTF10001234",
                        'contractNumber': "1234AB567891",
                        'description': "MYFSW1",
                        'isGovernment': false,
                    }
                ]

        Yields
        ------
        dict
            Dictionary containing information and result returned by the registration API call::

            {
                "serialNumber": "S424ENTF10001234",
                "description": "MYFSW1",
                "contractNumber": "1234AB567891",
                "status": 0,
                "message": "Success",
                "gbl_status": 0,
                "gbl_message": "Success",
            }
        """

        url = self.api_base + "products/register"

        asset_len = len(assets)

        if asset_len <= self.batch_reg_max:
            data = {"registrationUnits": assets} # TODO: Add support for registration without contract number, ideally with a warning for products that do not have contracts bundled (all but FortiGate?)

            r, r_sc, r_json = self._send_api_request(url, data=data)

            if r_json and r_sc in [200, 400]:
                if "assets" in r_json and r_json["assets"]:
                    for i, a in enumerate(r_json["assets"], 1):
                        if i == asset_len:
                            yield {
                                "serialNumber": a["serialNumber"],
                                "description": a["description"],
                                "contractNumber": a["contractNumber"],
                                "status": a["status"],
                                "message": a["message"],
                                "gbl_status": r_json["status"],
                                "gbl_message": r_json["message"],
                            }
                        else:
                            yield {
                                "serialNumber": a["serialNumber"],
                                "description": a["description"],
                                "contractNumber": a["contractNumber"],
                                "status": a["status"],
                                "message": a["message"],
                                "gbl_status": None,
                                "gbl_message": None,
                            }
                else:
                    yield {
                        "serialNumber": None,
                        "description": None,
                        "contractNumber": None,
                        "status": None,
                        "message": None,
                        "gbl_status": r_json["status"],
                        "gbl_message": r_json["message"],
                    }

            else:
                yield {
                    "serialNumber": None,
                    "description": None,
                    "contractNumber": None,
                    "status": None,
                    "message": None,
                    "gbl_status": 2,
                    "gbl_message": r_json["message"],
                }

    def registration_check(self, assets, check_type, check_period):
        """Checks if Fortinet assets are already registered, from a list of dictionaries containing asset information, using the FortiCare API.

        Parameters
        ----------
        assets : list
            List of dictionaries containing asset data::
            
                [
                    {
                        'serialNumber': "S424ENTF10001234",
                        'contractSku': "FC-10-S424E-247-02-36",
                        'contractNumber': "1234AB567891",
                        'description': "MYFSW1",
                        'isGovernment': false,
                    }
                ]

        check_type : str
            Contract pre-check type to use.
        check_period : str
            Contract pre-check period to use.

        Yields
        ------
        dict
            Dictionary containing True if registered, False if not registered, checked serial number string, description match bool, contract match bool, additional contracts list, and message string with the details::

            {
                "registered": True,
                "serialNumber": "S424ENTF10001234",
                "contractMatch": True,
                "descriptionMatch": True,
                "extraContracts": ["1234AB567892"],
                "message": "Serial number S424ENTF10001234 is already registered, description MYFSW1 matches, contract 1234AB567891 matches, additional contracts 1234AB567892 are also associated with this asset."
            }
        """

        date_today = date.today()
        if check_period == "3 months":
            check_date = date_today + relativedelta(months=3)
        elif check_period == "1 month":
            check_date = date_today + relativedelta(months=1)
        else:
            check_date = date_today + relativedelta(years=10)

        url = self.api_base + "products/list"

        for a in assets:
            sn_match = False
            desc_match = False
            api_desc = ""
            ctrct_match = False
            extra_ctrcts = []
            check_passed = True

            data = {"serialNumber": a["serialNumber"]}

            r, r_sc, r_json = self._send_api_request(url, data=data)

            if r_json and r_sc in [200]:
                if "assets" in r_json and r_json["assets"]:
                    if a["serialNumber"] == r_json["assets"][0]["serialNumber"]:
                        sn_match = True

                        if a["description"] == r_json["assets"][0]["description"]:
                            desc_match = True
                        else:
                            api_desc = r_json["assets"][0]["description"]

                        if (
                            "contracts" in r_json["assets"][0]
                            and r_json["assets"][0]["contracts"]
                        ):
                            for c in r_json["assets"][0]["contracts"]:

                                if a["contractNumber"] == c["contractNumber"]: # check if contract number is already registered against this asset
                                    ctrct_match = True
                                else:
                                    extra_ctrcts.append(c["contractNumber"])

                                if check_type == "any": # check all existing contracts for end dates within the check period
                                    for ent in r_json["assets"][0]["entitlements"]:
                                        end_date_str = ent["endDate"].split("T")[0].split("-")
                                        end_date = date(int(end_date_str[0]), int(end_date_str[1]), int(end_date_str[2]))

                                        if check_date < end_date: # check if the existing entitlement end date runs beyond the check period
                                            check_passed = False

                                elif check_type == "matching": # check existing contracts with matching SKU's for end dates within the check period
                                    if a["contractSku"] == c["sku"]:
                                        for term in c["terms"]:
                                            end_date_str = term["endDate"].split("T")[0].split("-")
                                            end_date = date(int(end_date_str[0]), int(end_date_str[1]), int(end_date_str[2]))

                                            if check_date < end_date: # check if the existing contract end date runs beyond the check period
                                                check_passed = False

            output_text = ""

            if sn_match == True:
                output_text = "Serial number {} is already registered".format(
                    a["serialNumber"]
                )

                if desc_match == True:
                    output_text += ", description {} matches".format(a["description"])
                else:
                    output_text += ", description {} does not match".format(api_desc)

                if ctrct_match == True:
                    output_text += ", contract {} matches".format(a["contractNumber"])

                    if extra_ctrcts:
                        output_text += ", additional contracts {} are also associated with this asset".format(
                            ", ".join(extra_ctrcts)
                        )
                else:
                    output_text += ", contract {} does not match".format(
                        a["contractNumber"]
                    )

                    if extra_ctrcts:
                        output_text += ", contracts {} are associated with this asset".format(
                            ", ".join(extra_ctrcts)
                        )

                if check_passed == False:
                    output_text += ", failed contracts pre-check"

                yield {
                    "registered": True,
                    "serialNumber": a["serialNumber"],
                    "contractMatch": ctrct_match,
                    "checkPassed": check_passed,
                    "descriptionMatch": desc_match,
                    "extraContracts": extra_ctrcts,
                    "message": output_text + ".\n",
                }

            else:

                yield {
                    "registered": False,
                    "serialNumber": a["serialNumber"],
                    "contractMatch": ctrct_match,
                    "checkPassed": check_passed,
                    "descriptionMatch": desc_match,
                    "extraContracts": extra_ctrcts,
                    "message": "SN not found.\n",
                }

    def add_contract(self, sn, ctrct):
        """Registers a contract against an existing asset using the FortiCare API.

        Parameters
        ----------
        sn : str
            Serial number of the asset to update.
        ctrct : str
            Contract registration code to register against the asset.

        Returns
        -------
        tuple
            Tuple of two values - successful/unsuccessful bool, message str::
            (
                True,
                "Success"
            )
        """

        # ? Should a check be added here to ensure the asset is already registered as this will register the asset if it does not exist already

        url = self.api_base + "products/register"

        data = {"registrationUnits": [{"serialNumber": sn, "contractNumber": ctrct}]}

        r, r_sc, r_json = self._send_api_request(url, data=data)

        if r_json and r_sc in [200]:
            if r_json["status"] == 0:
                return True, r_json["message"]  # Success

            else:
                return False, r_json["message"]  # API returned an error

        if r_json and "message" in r_json and r_json["message"]:
            return False, "API Error - {}".format(r_json["message"])

        return False, "API error"  # If we get here, API call probably failed

    def update_description(self, sn, desc):
        """Updates the description of an asset using the FortiCare API.

        Parameters
        ----------
        sn : str
            Serial number of the asset to update.
        desc : str
            New description for the asset.

        Returns
        -------
        bool
            True if successful, False if unsuccessful.
        """

        url = self.api_base + "products/description"

        data = {"serialNumber": sn, "description": desc}

        r, r_sc, r_json = self._send_api_request(url, data=data)

        if r_json and r_sc in [200]:
            if r_json["status"] == 0:
                return True, r_json["message"]  # Success

            else:
                return False, r_json["message"]  # API returned an error

        if r_json and "message" in r_json and r_json["message"]:
            return False, "API Error - {}".format(r_json["message"])

        return False, "API error"  # If we get here, API call probably failed


class FortiCRP:
    """A class to represent a Fortinet Contract Reader & Parser instance.


    Parameters
    ----------
    license_dir : str
        Path to the directory containing Fortinet contract ZIP archives.
    """

    def __init__(self, license_dir):
        """Constructs all the necessary attributes for the FortiCRP object.
        """

        from PyPDF2 import PdfFileReader
        from PyPDF2.utils import PdfReadError

        self.PdfFileReader = PdfFileReader
        self.PdfReadError = PdfReadError

        # Read file listing of license directory

        dir_file_list = os.listdir(license_dir)

        self.license_dir = license_dir

        self.license_zip_list = []

        for dir_file in dir_file_list:

            if dir_file.endswith(".zip"):

                self.license_zip_list.append(dir_file)

        # Static variables

        self.parsed_licenses = {}
        self.ignored_files_patterns = ["__MACOSX"]

        # Define regex patterns

        self.re_lic_regcode = (
            r"Registration Code\s\s\s:\s\s(.....-.....-.....-.....-......)"
        )

        self.re_supp_regcode = r"ContractRegistrationCode:(.+?)Support"

    def _parse_licenses(self):

        for zip_file in self.license_zip_list:

            contract_sku = zip_file.split("_")[0]

            if contract_sku not in self.parsed_licenses:
                self.parsed_licenses[contract_sku] = []

            zip_file_contents = ZipFile(self.license_dir + "/" + zip_file, "r")

            zip_file_contents_namelist = zip_file_contents.namelist()

            for ignored_pattern in self.ignored_files_patterns:
                for i, file_name in enumerate(zip_file_contents_namelist):
                    if ignored_pattern in file_name:
                        zip_file_contents_namelist.pop(i)

            for pdf_file in zip_file_contents_namelist:
                pdf_raw = zip_file_contents.open(pdf_file, "r")

                pdf_data = self.PdfFileReader(pdf_raw)

                pdf_text = ""

                for page in range(pdf_data.numPages):
                    page_data = pdf_data.getPage(page)
                    pdf_text += page_data.extractText()

                system = platform.system()
                if system == "Windows":
                    pdf_text = pdf_text.replace("\n", "")

                self.parsed_licenses[contract_sku].append(pdf_text)

    def get_parsed_licenses(self):
        """Parses Fortinet contract PDF files from ZIP archives in the provided license directory, returning a dictionary of full document contents.

        Returns
        -------
        dict
            Dictionary containing full contents of parsed PDF contract files::

            {
                "1234AB567891": "parsed files contents here..."
            }
        """

        if not self.parsed_licenses:
            self._parse_licenses()

        return self.parsed_licenses

    def get_registration_codes(self):
        """Parses Fortinet contract PDF files from ZIP archives in the provided license directory, returning valid license and support registration codes that are found.

        Yields
        ------
        tuple
            Tuple containing license or support registration code and associated contract SKU::

            (
                "FC-10-S424E-247-02-60",
                "1234AB567891"
            )
        """

        if not self.parsed_licenses:
            self._parse_licenses()

        for contract_sku, license_data in self.parsed_licenses.items():

            for data in license_data:

                reg_code = ""

                if not reg_code:

                    re_match = re.search(self.re_lic_regcode, data)

                    if re_match:
                        yield (re_match.group(1), contract_sku)

                if not reg_code:

                    re_match = re.search(self.re_supp_regcode, data)

                    if re_match:
                        yield (re_match.group(1), contract_sku)


def FCAT_UI():
    """Graphical User Interface (GUI) for the Fortinet Contract Automation Tool (FCAT).

    Fortinet Contract Automation Tool (FCAT) is a Python application to assist
    with bulk parsing of Fortinet contract ZIP files and the registration of
    Fortinet products and services using Fortinet's FortiCare API.

    Initialises the FCAT GUI.
    """

    import PySimpleGUI as sg
    import keyring

    fcat_key_name = "fcat.credentials"

    system = platform.system()
    if system == "Windows":
        from keyring.backends import Windows

        keyring.set_keyring(Windows.WinVaultKeyring())
    elif system == "Darwin":
        from keyring.backends import OS_X

        keyring.set_keyring(OS_X.Keyring())
    else:
        pass  # try autodiscovery on other platforms

    keyring_user = keyring.get_credential(fcat_key_name, "username")
    fcat_creds_user = keyring_user.password if keyring_user else None

    keyring_pass = keyring.get_credential(fcat_key_name, "password")
    fcat_creds_pass = keyring_pass.password if keyring_pass else None

    collapse_closed = "►"
    collapse_open = "▼"

    output_text = ""

    thread_key_start = "-THREADSTART-"
    thread_key_feedback = "-THREADFEEDBACK-"
    thread_key_end = "-THREADEND-"

    global thread_run

    def _csv_initialise(license_dir):

        timestamp = str(datetime.now().strftime("%Y%m%d_%H%M%S"))

        output_file = license_dir + "/licenses_" + timestamp + ".csv"

        with open(output_file, "w", newline="", encoding="utf-8",) as csv_file:

            csv_writer = csv.DictWriter(
                csv_file,
                fieldnames=[
                    "contract_sku",
                    "serial_number",
                    "registration_code",
                    "description",
                ],
                delimiter=",",
            )

            csv_writer.writeheader()

        return output_file

    def _csv_generate_empty(out_file):

        with open(out_file, "w", newline="", encoding="utf-8",) as csv_file:

            csv_writer = csv.DictWriter(
                csv_file,
                fieldnames=[
                    "contract_sku",
                    "serial_number",
                    "registration_code",
                    "description",
                ],
                delimiter=",",
            )

            csv_writer.writeheader()

        return out_file

    def _csv_write_contract(output_file, sku, rc):

        with open(output_file, "a", newline="", encoding="utf-8",) as csv_file:

            csv_writer = csv.DictWriter(
                csv_file,
                fieldnames=[
                    "contract_sku",
                    "serial_number",
                    "registration_code",
                    "description",
                ],
                delimiter=",",
            )

            line = {
                "contract_sku": sku,
                "serial_number": "",
                "registration_code": rc,
                "description": "",
            }

            csv_writer.writerow(line)

    def _parse_data_file(data_file):

        data_format = "nosku"  # "nosku" or "sku"

        if data_file.endswith(".xlsx"):
            from openpyxl import load_workbook

            wb = load_workbook(filename=data_file)
            ws = wb.active

            headers = []
            for cell in ws["1"]:
                headers.append(cell.value)

            if headers and headers[0] == "contract_sku":
                data_format = "sku"

            for row in ws.iter_rows(min_row=2):
                row_data = {}
                row_vals = []

                for cell in row:
                    row_vals.append(cell.value)

                if data_format == "nosku" and len(row_vals) >= 3:
                    row_data["contractSku"] = ""
                    row_data["serialNumber"] = row_vals[0].strip()
                    row_data["contractNumber"] = row_vals[1].strip()
                    row_data["description"] = row_vals[2].strip()
                elif data_format == "sku" and len(row_vals) >= 4:
                    row_data["contractSku"] = row_vals[0].strip()
                    row_data["serialNumber"] = row_vals[1].strip()
                    row_data["contractNumber"] = row_vals[2].strip()
                    row_data["description"] = row_vals[3].strip()

                yield row_data

        elif data_file.endswith(".csv"):
            import csv

            with open(data_file, "r", newline="", encoding="utf-8",) as csv_file:
                csv_reader = csv.reader(csv_file)

                headers = []
                for val in next(csv_reader, None):
                    headers.append(val)

                if headers and headers[0] == "contract_sku":
                    data_format = "sku"

                for row in csv_reader:

                    row_data = {}

                    if data_format == "nosku" and len(row) >= 3:
                        row_data["contractSku"] = ""
                        row_data["serialNumber"] = row[0].strip()
                        row_data["contractNumber"] = row[1].strip()
                        row_data["description"] = row[2].strip()
                    elif data_format == "sku" and len(row) >= 4:
                        row_data["contractSku"] = row[0].strip()
                        row_data["serialNumber"] = row[1].strip()
                        row_data["contractNumber"] = row[2].strip()
                        row_data["description"] = row[3].strip()

                    yield row_data

    def _process_reg_response(regd):

        output_text = ""
        regd_asset_counter = 0

        if regd["status"] == 0:
            output_text = (
                output_text
                + "Registered asset {} ({}) with contract number {}\n".format(
                    regd["serialNumber"], regd["description"], regd["contractNumber"],
                )
            )

            regd_asset_counter += 1

        elif regd["serialNumber"]:
            output_text = output_text + "Failed to register asset {}: {}\n".format(
                regd["serialNumber"], regd["message"],
            )

        if regd["gbl_status"] and regd["gbl_status"] != 0:
            if regd["gbl_message"] != "Failed":
                output_text = output_text + "API error: {}\n".format(
                    regd["gbl_message"],
                )

        return output_text, regd_asset_counter

    def _process_gui_registration(
        window,
        data_file,
        is_govt,
        check_type,
        check_period,
        fcat_creds_user,
        fcat_creds_pass,
        assets_buffer_size=5,
    ):

        if fcat_creds_user and fcat_creds_pass:

            fcapi = FortiCareAPI(fcat_creds_user, fcat_creds_pass)
            auth_success = fcapi.get_oauth_token()

            if auth_success:

                asset_counter = 0
                regd_asset_counter = 0

                if data_file:
                    window.write_event_value(
                        thread_key_feedback, "Parsing data file: {}\n".format(data_file)
                    )
                else:
                    window.write_event_value(
                        thread_key_feedback, "Parsing data file: no data file provided\n"
                    )

                if check_type in ["matching", "any"]:
                    window.write_event_value(
                        thread_key_feedback, "Contracts pre-check rule setting: {} contracts with over {} remaining\n".format(check_type, check_period)
                    )
                else:
                    window.write_event_value(
                        thread_key_feedback, "Contracts pre-check rule setting: existing contract check disabled\n"
                    )

                window.write_event_value(
                    thread_key_feedback,
                    "\n################## START PARSED ASSETS ##################\n",
                )

                assets_buffer = []

                for asset in _parse_data_file(data_file):

                    if all(
                        k in asset
                        for k in ("contractSku", "serialNumber", "contractNumber", "description")
                    ):

                        asset_data = {
                            "contractSku": asset["contractSku"],
                            "serialNumber": asset["serialNumber"],
                            "contractNumber": asset["contractNumber"],
                            "description": asset["description"],
                            "isGovernment": is_govt,
                        }

                        if len(assets_buffer) < assets_buffer_size:

                            for reg_chk in fcapi.registration_check([asset_data], check_type, check_period):

                                if reg_chk["registered"] == False:
                                    assets_buffer.append(asset_data)
                                else:
                                    window.write_event_value(
                                        thread_key_feedback, reg_chk["message"],
                                    )

                                    if reg_chk["descriptionMatch"] == False:
                                        upd_desc, ud_msg = fcapi.update_description(
                                            asset_data["serialNumber"],
                                            asset_data["description"],
                                        )

                                        if upd_desc:
                                            window.write_event_value(
                                                thread_key_feedback,
                                                "     └ Updated description to {} for {}.\n".format(
                                                    asset_data["description"],
                                                    asset_data["serialNumber"],
                                                ),
                                            )
                                        else:
                                            window.write_event_value(
                                                thread_key_feedback,
                                                "     └ Failed to update description to {} for {}: {}.\n".format(
                                                    asset_data["description"],
                                                    asset_data["serialNumber"],
                                                    ud_msg,
                                                ),
                                            )

                                    if reg_chk["checkPassed"] == True:

                                        if reg_chk["contractMatch"] == False:
                                            add_ctrct, ac_msg = fcapi.add_contract(
                                                asset_data["serialNumber"],
                                                asset_data["contractNumber"],
                                            )

                                            if add_ctrct:
                                                window.write_event_value(
                                                    thread_key_feedback,
                                                    "     └ Added contract {} to {}.\n".format(
                                                        asset_data["contractNumber"],
                                                        asset_data["serialNumber"],
                                                    ),
                                                )

                                            else:
                                                window.write_event_value(
                                                    thread_key_feedback,
                                                    "     └ Failed to add contract {} to {}: {}.\n".format(
                                                        asset_data["contractNumber"],
                                                        asset_data["serialNumber"],
                                                        ac_msg,
                                                    ),
                                                )

                                    else:
                                        window.write_event_value(
                                            thread_key_feedback,
                                            "     └ Failed existing contracts pre-check.\n"
                                        )

                        else:

                            for regd in fcapi.register_assets(assets_buffer):
                                (
                                    prr_output_text,
                                    prr_regd_asset_counter,
                                ) = _process_reg_response(regd)

                                window.write_event_value(
                                    thread_key_feedback, prr_output_text,
                                )

                                regd_asset_counter += prr_regd_asset_counter

                            assets_buffer = []

                            for reg_chk in fcapi.registration_check([asset_data]):

                                if reg_chk["registered"] == False:
                                    assets_buffer.append(asset_data)
                                else:
                                    window.write_event_value(
                                        thread_key_feedback, reg_chk["message"],
                                    )

                                    if reg_chk["descriptionMatch"] == False:
                                        upd_desc, ud_msg = fcapi.update_description(
                                            asset_data["serialNumber"],
                                            asset_data["description"],
                                        )

                                        if upd_desc:
                                            window.write_event_value(
                                                thread_key_feedback,
                                                "     └ Updated description to {} for {}.\n".format(
                                                    asset_data["description"],
                                                    asset_data["serialNumber"],
                                                ),
                                            )
                                        else:
                                            window.write_event_value(
                                                thread_key_feedback,
                                                "     └ Failed to update description to {} for {}: {}.\n".format(
                                                    asset_data["description"],
                                                    asset_data["serialNumber"],
                                                    ud_msg,
                                                ),
                                            )

                                    if reg_chk["checkPassed"] == True:

                                        if reg_chk["contractMatch"] == False:
                                            add_ctrct, ac_msg = fcapi.add_contract(
                                                asset_data["serialNumber"],
                                                asset_data["contractNumber"],
                                            )

                                            if add_ctrct:
                                                window.write_event_value(
                                                    thread_key_feedback,
                                                    "     └ Added contract {} to {}.\n".format(
                                                        asset_data["contractNumber"],
                                                        asset_data["serialNumber"],
                                                    ),
                                                )

                                            else:
                                                window.write_event_value(
                                                    thread_key_feedback,
                                                    "     └ Failed to add contract {} to {}: {}.\n".format(
                                                        asset_data["contractNumber"],
                                                        asset_data["serialNumber"],
                                                        ac_msg,
                                                    ),
                                                )

                                    else:
                                        window.write_event_value(
                                            thread_key_feedback,
                                            "     └ Failed existing contracts pre-check.\n"
                                        )

                        asset_counter += 1

                    if thread_run == False:
                        break

                if thread_run == True and len(assets_buffer) != 0:

                    for regd in fcapi.register_assets(assets_buffer):
                        (
                            prr_output_text,
                            prr_regd_asset_counter,
                        ) = _process_reg_response(regd)

                        window.write_event_value(
                            thread_key_feedback, prr_output_text,
                        )

                        regd_asset_counter += prr_regd_asset_counter

                window.write_event_value(
                    thread_key_feedback,
                    "################## END PARSED ASSETS ##################\n\n",
                )

                if thread_run == False:
                    window.write_event_value(
                        thread_key_feedback, "Terminated thread at user request.\n\n",
                    )

                window.write_event_value(
                    thread_key_feedback,
                    "Parsing complete - {} assets found, {} assets registered.\n".format(
                        asset_counter, regd_asset_counter,
                    ),
                )

                fcapi.revoke_oauth_token()

            else:
                output_text = "Authentication to FortiCare API failed.\n"
                window["OUTPUT"].update(value=output_text, autoscroll=True)

        else:
            output_text = "Need to provide FortiCare API credentials before assets can be registered.\n"
            window["OUTPUT"].update(value=output_text, autoscroll=True)

    def _process_gui_details(
        window,
        serial_numbers,
        fcat_creds_user,
        fcat_creds_pass,
    ):

        if fcat_creds_user and fcat_creds_pass:

            fcapi = FortiCareAPI(fcat_creds_user, fcat_creds_pass)
            auth_success = fcapi.get_oauth_token()

            if auth_success:

                window.write_event_value(
                    thread_key_feedback,
                    "\n################## START ASSET DETAILS ##################\n",
                )

                sn_list = serial_numbers.split("\n")

                for sn in sn_list:
                    if sn:
                        output_text = "{}\n".format(sn)
                        window.write_event_value(thread_key_feedback, output_text)

                    if thread_run == False:
                        break

                window.write_event_value(
                    thread_key_feedback,
                    "################## END ASSET DETAILS ##################\n\n",
                )

                if thread_run == False:
                    window.write_event_value(
                        thread_key_feedback, "Terminated thread at user request.\n\n",
                    )

                sn_list_count = len(sn_list)

                window.write_event_value(
                    thread_key_feedback,
                    "Request complete - {} assets details found.\n".format(sn_list_count),
                )

                fcapi.revoke_oauth_token()

            else:
                output_text = "Authentication to FortiCare API failed.\n"
                window["OUTPUT"].update(value=output_text, autoscroll=True)

        else:
            output_text = "Need to provide FortiCare API credentials before assets can be registered.\n"
            window["OUTPUT"].update(value=output_text, autoscroll=True)

    def _save_output_to_file(out_file, out_text):

        with open(out_file, "w") as f:
            f.write(out_text)

    def _make_window(theme="Reddit"):

        sg.theme(theme)

        font = ("Helvetica", 16)
        sg.set_options(font=font)

        elements = [[sg.T("")]]

        tab_contract_parser = [
            [sg.T("")],
            [
                sg.Text(
                    "Select a directory containing Fortinet contract ZIP files to parse...",
                    size=(120, 1),
                    justification="center",
                )
            ],
            [
                sg.Input(key="LICDIR2", change_submits=True,),
                sg.FolderBrowse(key="LICDIR"),
            ],
            [sg.Button("Parse Contracts")],
            [
                sg.InputText(
                    "",
                    do_not_clear=False,
                    visible=False,
                    key="Generate Template",
                    enable_events=True,
                ),
                sg.FileSaveAs(
                    "Generate Empty Template",
                    file_types=[("CSV", ".csv")],
                    initial_folder=None,
                    default_extension=".csv",
                    button_color=("orange", "white"),
                ),
            ],
            [sg.T("")],
        ]

        tab_asset_registration = [
            [sg.T("")],
            [
                sg.Text(
                    "Select a CSV or XLSX data file containing assets to register...",
                    size=(120, 1),
                    justification="center",
                )
            ],
            [
                sg.Input(key="DATAFILE", change_submits=True,),
                sg.FileBrowse(file_types=(("CSV", "*.csv"), ("XLSX", "*.xlsx"))),
            ],
            [
                sg.Checkbox(
                    "Is this a Government organisation? (leave unchecked for no)",
                    default=False,
                    key="ISGOVT",
                )
            ],
            [
                sg.Button("Stop Registration", button_color=("red", "white")),
                sg.Button("Register Assets"),
            ],
            [sg.T("")],
        ]

        tab_asset_details = [
            [sg.T("")],
            [
                sg.Text(
                    "Enter serial numbers seperated by new lines to get details...",
                    size=(120, 1),
                    justification="center",
                )
            ],
            [sg.Multiline("", key="SERIALS", size=(50, 3))],

            [
                sg.Button("Stop Details", button_color=("red", "white")),
                sg.Button("Get Details"),
            ],
            [sg.T("")],
        ]

        if fcat_creds_user and fcat_creds_pass:
            user_placeholder = fcat_creds_user
            pass_placeholder = fcat_creds_pass
        else:
            user_placeholder = ""
            pass_placeholder = ""

        check_type = [
            "matching",
            "any",
            "none"
        ]

        check_period = [
            "3 months",
            "1 month"
        ]

        tab_settings = [
            [sg.T("")],
            [
                sg.Text("Username:", size=(8, 1)),
                sg.InputText(user_placeholder, key="USERNAME"),
                sg.Button("Update Credentials"),
            ],
            [
                sg.Text("Password:", size=(8, 1)),
                sg.InputText(pass_placeholder, key="PASSWORD", password_char="*"),
                sg.Button("Delete Credentials", button_color=("red", "white")),
            ],
            [
                sg.Text("Contract Check Type:", size=(18, 1)),
                sg.Combo(check_type, check_type[0], key="CHECKTYPE", size=(50, 1)),
            ],
            [
                sg.Text("Contract Check Period:", size=(18, 1)),
                sg.Combo(check_period, check_period[0], key="CHECKPERIOD", size=(50, 1)),
            ],

            [sg.T("")],
        ]

        elements += [
            [
                sg.TabGroup(
                    [
                        [
                            sg.Tab(
                                "Contract Parser",
                                tab_contract_parser,
                                element_justification="center",
                            ),
                        ],
                        [
                            sg.Tab(
                                "Asset Registration",
                                tab_asset_registration,
                                element_justification="center",
                            ),
                        ],
                        [
                            sg.Tab(
                                "Asset Details",
                                tab_asset_details,
                                element_justification="center",
                            ),
                        ],
                        [
                            sg.Tab(
                                "Settings",
                                tab_settings,
                                element_justification="center",
                            ),
                        ],
                    ],
                    key="FUNCTABS",
                    change_submits=True,
                )
            ]
        ]

        elements += [
            [sg.T("")],
            [sg.Text("Output")],
            [sg.Multiline("", key="OUTPUT", size=(120, 15), disabled=True)],
            [
                sg.Button("Clear Output", button_color=("orange", "white")),
                sg.InputText(
                    "",
                    do_not_clear=False,
                    visible=False,
                    key="Save Output",
                    enable_events=True,
                ),
                sg.FileSaveAs(
                    "Save Output to File",
                    file_types=[("TXT", ".txt")],
                    initial_folder=None,
                    default_extension=".txt",
                ),
            ],
            [sg.T("")],
        ]

        menu_def = [
            ["FCAT", ["About", "Exit"]],
            ["Theme", sg.theme_list()],
        ]

        layout = [
            [sg.Menu(menu_def, key="-MENU-",)],
            [sg.VPush()],
            [
                sg.Push(),
                sg.Column(elements, element_justification="center"),
                sg.Push(),
            ],
            [sg.VPush()],
        ]

        return sg.Window("Fortinet Contract Automation Tool (FCAT)", layout, resizable=True)

    # START GUI

    window = _make_window()

    while True:
        event, values = window.read()

        if event in (sg.WIN_CLOSED, "Exit"):
            break

        elif event == thread_key_feedback:

            output_text = output_text + values[thread_key_feedback]
            window["OUTPUT"].update(value=output_text, autoscroll=True)

        elif event == "Parse Contracts":

            license_dir = values["LICDIR"]

            written_contracts = []

            try:
                output_file = _csv_initialise(license_dir)

                fcrp = FortiCRP(license_dir)

                if license_dir:
                    output_text = "Parsing directory: {}\n".format(license_dir)
                else:
                    output_text = "Parsing directory: no directory provided\n"

                output_text = (
                    output_text
                    + "\n################## START PARSED LICENSES ##################\n"
                )

                for rc, sku in fcrp.get_registration_codes():
                    if rc not in written_contracts:
                        _csv_write_contract(output_file, sku, rc)
                        written_contracts.append(rc)
                        output_text = (
                            output_text
                            + "Parsed registration code {} for SKU {}\n".format(rc, sku)
                        )
                        window["OUTPUT"].update(value=output_text, autoscroll=True)
                    else:
                        output_text = (
                            output_text
                            + "WARNING: Duplicate registration code {} for SKU {}\n".format(
                                rc, sku
                            )
                        )
                        window["OUTPUT"].update(value=output_text, autoscroll=True)

            except Exception as e:
                sg.Popup(
                    "Failed to parse contract files, see debug window for details."
                )
                tb = format_exc(e)
                sg.Print(tb, keep_on_top=True)

            output_text = (
                output_text
                + "################## END PARSED LICENSES ##################\n\n"
            )
            window["OUTPUT"].update(value=output_text, autoscroll=True)

            output_text = output_text + "Parser output: {}".format(output_file)
            window["OUTPUT"].update(value=output_text, autoscroll=True)

        elif event == "Generate Template":

            out_file = values["Generate Template"]

            try:
                _csv_generate_empty(out_file)
                sg.Popup(
                    "Successfully generated template contracts file: {}".format(
                        out_file
                    )
                )

            except Exception as e:
                sg.Popup(
                    "Failed to generate template contracts file, see debug window for details."
                )
                tb = format_exc(e)
                sg.Print(tb, keep_on_top=True)

        elif event == "Register Assets":

            data_file = values["DATAFILE"]
            is_govt = values["ISGOVT"]
            check_type = values["CHECKTYPE"]
            check_period = values["CHECKPERIOD"]

            output_text = "Starting registration thread\n"
            window["OUTPUT"].update(value=output_text, autoscroll=True)

            thread_run = True

            window.start_thread(
                lambda: _process_gui_registration(
                    window, data_file, is_govt, check_type, check_period, fcat_creds_user, fcat_creds_pass,
                ),
                thread_key_end,
            )

        elif event == "Stop Registration":

            if thread_run:
                thread_run = False
                sg.Popup("Stopping registration thread on next loop.")
            else:
                sg.Popup("Registration thread is not running.")

        elif event == "Get Details":

            serial_numbers = values["SERIALS"]

            output_text = "Starting get details thread\n"
            window["OUTPUT"].update(value=output_text, autoscroll=True)

            thread_run = True

            window.start_thread(
                lambda: _process_gui_details(
                    window, serial_numbers, fcat_creds_user, fcat_creds_pass,
                ),
                thread_key_end,
            )

        elif event == "Stop Details":

            if thread_run:
                thread_run = False
                sg.Popup("Stopping details thread on next loop.")
            else:
                sg.Popup("Details thread is not running.")

        elif event == "Clear Output":

            window["OUTPUT"].update(value="")

        elif event == "Save Output":

            out_file = values["Save Output"]
            out_text = values["OUTPUT"]

            try:
                if out_file and out_text:
                    _save_output_to_file(out_file, out_text)
                    sg.Popup("Output saved successfully to file {}".format(out_file))
                else:
                    sg.Popup("No output to save.")

            except Exception as e:
                sg.Popup("Failed to save output to file, see debug window for details.")
                tb = format_exc(e)
                sg.Print(tb, keep_on_top=True)

        elif event == "Update Credentials":

            fcat_creds_user = values["USERNAME"]
            fcat_creds_pass = values["PASSWORD"]

            try:
                keyring.set_password(fcat_key_name, "username", fcat_creds_user)
                keyring.set_password(fcat_key_name, "password", fcat_creds_pass)
                sg.Popup("Credentials saved successfully.")

            except Exception as e:
                sg.Popup("Failed to save credentials.")
                tb = format_exc(e)
                sg.Print(tb, keep_on_top=True)

        elif event == "Delete Credentials":

            try:
                keyring.delete_password(fcat_key_name, "username")
                keyring.delete_password(fcat_key_name, "password")
                window["USERNAME"].update(value="")
                window["PASSWORD"].update(value="")
                sg.Popup("Credentials deleted successfully.")

            except Exception as e:
                sg.Popup("Failed to delete credentials, see debug window for details.")
                tb = format_exc(e)
                sg.Print(tb, keep_on_top=True)

        elif event == "About":

            sg.popup(
                __title__,
                __description__,
                "Version: {}".format(__version__),
                "Author: {}".format(__author__),
                title="About",
            )

        elif event in sg.theme_list():

            window.close()
            window = _make_window(theme=event)

    window.close()


if __name__ == "__main__":
    try:
        FCAT_UI()
    except Exception as e:
        import PySimpleGUI as sg

        sg.theme("Reddit")

        font = ("Helvetica", 16)
        sg.set_options(font=font)

        tb = format_exc()

        sg.Print(tb, keep_on_top=True)
        sg.popup_error(
            "ERROR",
            "See the debug window for details.",
            "Please share these details with the application developer.",
            keep_on_top=True,
        )

